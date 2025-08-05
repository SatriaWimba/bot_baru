import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def create_calendar_excel(year=2025, filename='kalender_2025_blackbox.xlsx'):
    # Membuat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Kalender 2025"
    
    # Style untuk header
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    # Menulis judul tahun
    ws.merge_cells('A1:G1')
    year_cell = ws['A1']
    year_cell.value = f'KALENDER TAHUN {year}'
    year_cell.font = Font(bold=True, size=16)
    year_cell.alignment = Alignment(horizontal='center')
    
    # Membuat kalender untuk setiap bulan
    row_start = 3
    for month in range(1, 13):
        # Menulis nama bulan
        month_name = calendar.month_name[month]
        ws.merge_cells(start_row=row_start, start_column=1, end_row=row_start, end_column=7)
        month_cell = ws.cell(row=row_start, column=1)
        month_cell.value = month_name.upper()
        month_cell.font = Font(bold=True, size=14)
        month_cell.alignment = Alignment(horizontal='center')
        row_start += 1
        
        # Header hari
        days = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        for i, day in enumerate(days, 1):
            cell = ws.cell(row=row_start, column=i)
            cell.value = day[:3]  # Singkatan hari
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Mengisi tanggal
        month_days = calendar.monthcalendar(year, month)
        row_start += 1
        for week in month_days:
            for day in range(7):
                date = week[day]
                cell = ws.cell(row=row_start, column=day+1)
                if date != 0:
                    cell.value = date
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            row_start += 1
        
        row_start += 2  # Spasi antar bulan
    
    # Mengatur lebar kolom
    for col in 'ABCDEFG':
        ws.column_dimensions[col].width = 12
    
    # Menyimpan file
    wb.save(filename)
    print(f'File {filename} berhasil dibuat!')

# Menjalankan fungsi
create_calendar_excel()
