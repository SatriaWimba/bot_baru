import calendar
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def buat_kalender_excel(tahun, nama_file):
    wb = Workbook()
    wb.remove(wb.active)  # Hapus sheet default

    hari = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

    for bulan in range(1, 13):
        sheet = wb.create_sheet(calendar.month_name[bulan])
        cal = calendar.monthcalendar(tahun, bulan)

        # Judul bulan
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        cell_judul = sheet.cell(row=1, column=1)
        cell_judul.value = f"{calendar.month_name[bulan]} {tahun}"
        cell_judul.font = Font(size=14, bold=True)
        cell_judul.alignment = Alignment(horizontal='center')

        # Header hari
        for i, nama_hari in enumerate(hari):
            sheet.cell(row=2, column=i+1).value = nama_hari
            sheet.cell(row=2, column=i+1).font = Font(bold=True)
            sheet.cell(row=2, column=i+1).alignment = Alignment(horizontal='center')

        # Isi tanggal
        for minggu_ke, minggu in enumerate(cal):
            for i, tanggal in enumerate(minggu):
                cell = sheet.cell(row=3 + minggu_ke, column=i+1)
                cell.value = tanggal if tanggal != 0 else ""
                cell.alignment = Alignment(horizontal='center')

        # Lebar kolom
        for col in range(1, 8):
            sheet.column_dimensions[chr(64 + col)].width = 10

    wb.save(nama_file)
    print(f"Kalender {tahun} berhasil disimpan sebagai {nama_file}")

# Panggil fungsi
buat_kalender_excel(2025, "kalender_2025_chatgpt.xlsx")
