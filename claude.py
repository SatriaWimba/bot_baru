import calendar
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def tampilkan_kalender_2025():
    """
    Menampilkan kalender lengkap untuk tahun 2025
    """
    tahun = 2025
    
    print("=" * 50)
    print(f"        KALENDER TAHUN {tahun}")
    print("=" * 50)
    
    # Nama bulan dalam bahasa Indonesia
    nama_bulan = [
        "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember"
    ]
    
    # Loop untuk setiap bulan
    for bulan in range(1, 13):
        print(f"\n{nama_bulan[bulan]} {tahun}".center(20))
        print("-" * 20)
        
        # Menampilkan header hari
        print("Sen Sel Rab Kam Jum Sab Min")
        
        # Mendapatkan kalender bulan
        cal = calendar.monthcalendar(tahun, bulan)
        
        # Menampilkan tanggal
        for minggu in cal:
            for hari in minggu:
                if hari == 0:
                    print("   ", end=" ")
                else:
                    print(f"{hari:2d}", end="  ")
            print()  # Baris baru setelah setiap minggu

def info_tahun_2025():
    """
    Menampilkan informasi tambahan tentang tahun 2025
    """
    print("\n" + "=" * 50)
    print("INFORMASI TAHUN 2025")
    print("=" * 50)
    
    # Cek apakah tahun kabisat
    if calendar.isleap(2025):
        print("✓ Tahun 2025 adalah tahun kabisat (366 hari)")
    else:
        print("✗ Tahun 2025 bukan tahun kabisat (365 hari)")
    
    # Hari pertama tahun
    hari_pertama = calendar.weekday(2025, 1, 1)
    nama_hari = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    print(f"• 1 Januari 2025 jatuh pada hari: {nama_hari[hari_pertama]}")
    
    # Hari terakhir tahun
    hari_terakhir = calendar.weekday(2025, 12, 31)
    print(f"• 31 Desember 2025 jatuh pada hari: {nama_hari[hari_terakhir]}")
    
    # Jumlah hari dalam setahun
    total_hari = 366 if calendar.isleap(2025) else 365
    print(f"• Total hari dalam tahun 2025: {total_hari} hari")

def cari_tanggal_khusus():
    """
    Mencari hari untuk tanggal-tanggal penting di tahun 2025
    """
    print("\n" + "=" * 50)
    print("HARI-HARI PENTING TAHUN 2025")
    print("=" * 50)
    
    tanggal_penting = [
        (1, 1, "Tahun Baru"),
        (2, 14, "Hari Kasih Sayang (Valentine)"),
        (8, 17, "Hari Kemerdekaan Indonesia"),
        (12, 25, "Hari Natal"),
        (12, 31, "Malam Tahun Baru")
    ]
    
    nama_hari = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
    
    for bulan, tanggal, acara in tanggal_penting:
        hari = calendar.weekday(2025, bulan, tanggal)
        print(f"• {tanggal:2d}/{bulan:2d}/2025 ({acara}): {nama_hari[hari]}")

def export_ke_excel():
    """
    Export kalender 2025 ke file Excel dengan format yang rapi
    """
    print("\nMemproses export ke Excel...")
    
    try:
        # Membuat workbook baru
        wb = Workbook()
        
        # Nama bulan dalam bahasa Indonesia
        nama_bulan = [
            "Januari", "Februari", "Maret", "April", "Mei", "Juni",
            "Juli", "Agustus", "September", "Oktober", "November", "Desember"
        ]
        
        # Menghapus sheet default
        wb.remove(wb.active)
        
        # Style untuk header
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Style untuk tanggal
        date_font = Font(size=10)
        date_alignment = Alignment(horizontal="center", vertical="center")
        weekend_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Membuat sheet untuk setiap bulan
        for bulan_idx, bulan_nama in enumerate(nama_bulan, 1):
            ws = wb.create_sheet(title=bulan_nama)
            
            # Header bulan dan tahun
            ws.merge_cells('A1:G1')
            ws['A1'] = f"{bulan_nama} 2025"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Header hari
            hari_header = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
            for col, hari in enumerate(hari_header, 1):
                cell = ws.cell(row=3, column=col, value=hari)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Mendapatkan kalender bulan
            cal = calendar.monthcalendar(2025, bulan_idx)
            
            # Mengisi tanggal
            for row_idx, minggu in enumerate(cal, 4):
                for col_idx, hari in enumerate(minggu, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = date_alignment
                    cell.font = date_font
                    
                    if hari != 0:
                        cell.value = hari
                        # Highlight weekend (Sabtu dan Minggu)
                        if col_idx in [6, 7]:  # Sabtu dan Minggu
                            cell.fill = weekend_fill
            
            # Mengatur lebar kolom
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 12
            
            # Mengatur tinggi baris
            for row in range(3, len(cal) + 4):
                ws.row_dimensions[row].height = 25
        
        # Membuat sheet ringkasan tahun
        summary_ws = wb.create_sheet(title="Ringkasan 2025", index=0)
        
        # Header ringkasan
        summary_ws['A1'] = "RINGKASAN KALENDER 2025"
        summary_ws['A1'].font = Font(bold=True, size=16)
        summary_ws.merge_cells('A1:C1')
        summary_ws['A1'].alignment = Alignment(horizontal="center")
        
        # Informasi tahun
        info_data = [
            ["Tahun", 2025],
            ["Jenis Tahun", "Bukan Tahun Kabisat"],
            ["Total Hari", 365],
            ["1 Januari 2025", calendar.day_name[calendar.weekday(2025, 1, 1)]],
            ["31 Desember 2025", calendar.day_name[calendar.weekday(2025, 12, 31)]]
        ]
        
        for row_idx, (label, value) in enumerate(info_data, 3):
            summary_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row_idx, column=2, value=value)
        
        # Hari-hari penting
        summary_ws['A9'] = "HARI-HARI PENTING"
        summary_ws['A9'].font = Font(bold=True, size=14)
        summary_ws.merge_cells('A9:D9')
        
        tanggal_penting = [
            (1, 1, "Tahun Baru"),
            (2, 14, "Hari Kasih Sayang"),
            (8, 17, "Hari Kemerdekaan RI"),
            (12, 25, "Hari Natal"),
            (12, 31, "Malam Tahun Baru")
        ]
        
        # Header tabel hari penting
        headers = ["Tanggal", "Bulan", "Acara", "Hari"]
        for col_idx, header in enumerate(headers, 1):
            cell = summary_ws.cell(row=11, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data hari penting
        for row_idx, (bulan, tanggal, acara) in enumerate(tanggal_penting, 12):
            hari = calendar.day_name[calendar.weekday(2025, bulan, tanggal)]
            summary_ws.cell(row=row_idx, column=1, value=tanggal)
            summary_ws.cell(row=row_idx, column=2, value=nama_bulan[bulan-1])
            summary_ws.cell(row=row_idx, column=3, value=acara)
            summary_ws.cell(row=row_idx, column=4, value=hari)
        
        # Mengatur lebar kolom untuk ringkasan
        column_widths = [12, 15, 25, 12]
        for col_idx, width in enumerate(column_widths, 1):
            summary_ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Menyimpan file
        filename = "Kalender_2025.xlsx"
        wb.save(filename)
        
        print(f"✅ Kalender berhasil di-export ke file: {filename}")
        print(f"📁 Lokasi file: {os.path.abspath(filename)}")
        print("\nFile Excel berisi:")
        print("• Sheet Ringkasan 2025 (informasi umum dan hari penting)")
        print("• 12 Sheet untuk setiap bulan dengan kalender lengkap")
        print("• Format yang rapi dengan highlight weekend")
        
    except ImportError:
        print("❌ Error: Library yang diperlukan tidak tersedia!")
        print("Silakan install library berikut:")
        print("pip install pandas openpyxl")
        
    except Exception as e:
        print(f"❌ Error saat export: {str(e)}")

def export_kalender_sederhana():
    """
    Export kalender dalam format CSV sederhana jika Excel tidak bisa
    """
    print("\nMemproses export ke CSV...")
    
    try:
        # Nama bulan
        nama_bulan = [
            "Januari", "Februari", "Maret", "April", "Mei", "Juni",
            "Juli", "Agustus", "September", "Oktober", "November", "Desember"
        ]
        
        # Membuat data untuk semua bulan
        all_data = []
        
        for bulan_idx in range(1, 13):
            bulan_nama = nama_bulan[bulan_idx - 1]
            cal = calendar.monthcalendar(2025, bulan_idx)
            
            # Header bulan
            all_data.append([f"\n{bulan_nama} 2025", "", "", "", "", "", ""])
            all_data.append(["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"])
            
            # Data kalender
            for minggu in cal:
                row = []
                for hari in minggu:
                    if hari == 0:
                        row.append("")
                    else:
                        row.append(str(hari))
                all_data.append(row)
            
            # Spasi antar bulan
            all_data.append(["", "", "", "", "", "", ""])
        
        # Menyimpan ke CSV
        import csv
        filename = "Kalender_2025.csv"
        with open(filename, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerows(all_data)
        
        print(f"✅ Kalender berhasil di-export ke file: {filename}")
        print(f"📁 Lokasi file: {os.path.abspath(filename)}")
        
    except Exception as e:
        print(f"❌ Error saat export: {str(e)}")

def menu_export():
    """
    Menu khusus untuk export
    """
    print("\n" + "=" * 50)
    print("MENU EXPORT KALENDER")
    print("=" * 50)
    print("1. Export ke Excel (.xlsx) - Format lengkap dan rapi")
    print("2. Export ke CSV (.csv) - Format sederhana")
    print("3. Kembali ke menu utama")
    
    pilihan = input("\nPilih format export (1-3): ").strip()
    
    if pilihan == "1":
        export_ke_excel()
    elif pilihan == "2":
        export_kalender_sederhana()
    elif pilihan == "3":
        return
    else:
        print("Pilihan tidak valid!")
        menu_export()
    """
    Menu interaktif untuk berbagai pilihan kalender
    """
    while True:
        print("\n" + "=" * 50)
        print("MENU KALENDER 2025")
        print("=" * 50)
        print("1. Tampilkan kalender lengkap tahun 2025")
        print("2. Tampilkan kalender bulan tertentu")
        print("3. Informasi tahun 2025")
        print("4. Hari-hari penting")
        print("5. Cari hari untuk tanggal tertentu")
        print("6. Keluar")
        
        pilihan = input("\nPilih menu (1-6): ").strip()
        
        if pilihan == "1":
            tampilkan_kalender_2025()
        
        elif pilihan == "2":
            try:
                bulan = int(input("Masukkan nomor bulan (1-12): "))
                if 1 <= bulan <= 12:
                    nama_bulan = [
                        "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
                        "Juli", "Agustus", "September", "Oktober", "November", "Desember"
                    ]
                    print(f"\n{nama_bulan[bulan]} 2025")
                    print(calendar.month(2025, bulan))
                else:
                    print("Nomor bulan harus antara 1-12!")
            except ValueError:
                print("Masukkan nomor yang valid!")
        
        elif pilihan == "3":
            info_tahun_2025()
        
        elif pilihan == "4":
            cari_tanggal_khusus()
        
        elif pilihan == "5":
            try:
                tanggal = int(input("Masukkan tanggal (1-31): "))
                bulan = int(input("Masukkan bulan (1-12): "))
                
                if 1 <= bulan <= 12:
                    # Cek apakah tanggal valid untuk bulan tersebut
                    max_hari = calendar.monthrange(2025, bulan)[1]
                    if 1 <= tanggal <= max_hari:
                        hari = calendar.weekday(2025, bulan, tanggal)
                        nama_hari = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
                        print(f"\n{tanggal}/{bulan}/2025 jatuh pada hari: {nama_hari[hari]}")
                    else:
                        print(f"Tanggal tidak valid! Bulan {bulan} hanya memiliki {max_hari} hari.")
                else:
                    print("Nomor bulan harus antara 1-12!")
            except ValueError:
                print("Masukkan nomor yang valid!")
        
        elif pilihan == "6":
            print("Terima kasih telah menggunakan Kalender 2025!")
            break
        
        else:
            print("Pilihan tidak valid! Silakan pilih 1-6.")

# Program utama
if __name__ == "__main__":
    print("🗓️  SELAMAT DATANG DI PROGRAM KALENDER 2025  🗓️")
    menu_interaktif()